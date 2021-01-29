from openpyxl import load_workbook

wb = load_workbook(r'test.xlsx')
ws = wb.active

row_i = 0
for a in range(0,10):
    row_i = row_i + 1
    ws['A%d' % row_i] =  row_i
wb.save('test.xlsx')